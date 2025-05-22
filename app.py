from flask import Flask, render_template, request, send_file
import sys
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime, timedelta
import random
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (NoSuchElementException, TimeoutException, WebDriverException)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
import threading
from queue import Queue
import os
import queue
import logging
no_result_dates = set()

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = '.'

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger(__name__)


def random_delay(min_sec=1, max_sec=2):
    time.sleep(random.uniform(min_sec, max_sec))


def human_like_interaction(driver):
    try:
        action = ActionChains(driver)
        for _ in range(random.randint(1, 3)):
            try:
                action.move_by_offset(0, 0).perform()
            except Exception as e:
                logger.warning(f"❌ Mouse move failed: {str(e).splitlines()[0]}")
            time.sleep(random.uniform(0.1, 0.3))

        if random.random() > 0.3:
            scroll_amount = random.randint(200, 600)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            time.sleep(random.uniform(0.5, 1.5))
    except (WebDriverException, Exception) as e:
        logger.warning(f"❌ Interaction simulation failed: {str(e).splitlines()[0]}")


def setup_driver():
    options = uc.ChromeOptions()
    buster_extension_path = r"C:\Users\ASUS TUF\AppData\Local\Google\Chrome\User Data\Default\Extensions\mpbjkejclgfgadiemmefgebjfooflfhl\3.1.0_0"

    if not os.path.exists(buster_extension_path):
        logger.error(f"❌ Extension path does not exist: {buster_extension_path}")
        sys.exit(1)

    options.add_argument(f"--load-extension={buster_extension_path}")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")
    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
    ]
    options.add_argument(f"user-agent={random.choice(user_agents)}")
    # options.add_argument(f"--window-size={random.randint(1000, 1400)},{random.randint(800, 1200)}")

    try:
        driver = uc.Chrome(
            options=options,
            headless=False,
            use_subprocess=True
        )

        driver.set_window_size(900, 450)

        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                window.navigator.chrome = {
                    runtime: {},
                };
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3]
                });
            """
        })
        return driver
    except Exception as e:
        logger.error(f"❌ Failed to initialize WebDriver: {str(e).splitlines()[0]}")
        sys.exit(1)


def handle_possible_blocking(driver, current_url):
    blocking_indicators = [
        "//div[contains(text(), 'Access Denied')]",
        "//div[contains(text(), 'Checking your browser')]",
        "//div[contains(text(), 'Please verify you are a human')]",
        "//iframe[contains(@title, 'recaptcha')]",
        "//div[contains(@class, 'cf-challenge')]"
    ]

    for indicator in blocking_indicators:
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, indicator)))
            logger.info("⚠️ Human verification detected. Clearing cookies and retrying after 1 minute...")
            driver.delete_all_cookies()
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")
            time.sleep(60)
            driver.quit()
            driver = setup_driver()
            driver.get(current_url)
            human_like_interaction(driver)
            random_delay(1, 2)
            return True, driver
        except TimeoutException:
            continue
    return False, driver


def scrape_flight_data_interval(driver_queue, results_queue, search_params, start_date):
    driver = None
    try:
        driver = driver_queue.get()
        nights = int(search_params['nights'])
        end_date_interval = start_date + timedelta(days=nights)
        date_from_str = start_date.strftime('%Y-%m-%d')
        date_to_str = end_date_interval.strftime('%Y-%m-%d')
        departure_airport = search_params['departure_airport']
        arrival_airport = search_params['arrival_airport']
        stops = search_params['stops']
        flight_hours = int(search_params['flight_hours'])
        country = search_params.get('country', 'USA')
        departure_airport_optional = search_params.get('departure_airport_optional')
        arrival_airport_optional = search_params.get('arrival_airport_optional')

        stops_param = ""
        if stops:
            stops_list = []
            if '0' in stops:
                stops_list.append("0")
            for stop in stops:
                if stop.isdigit() and stop != '0':
                    stops_list.append(stop)
            if stops_list:
                stops_param = ";stops=" + ",".join(stops_list)

        base_url = "https://www.kayak.com/flights"
        if country == 'Canada':
            base_url = "https://www.ca.kayak.com/flights"

        url = f"{base_url}/{departure_airport}-{arrival_airport}/{date_from_str}/{date_to_str}/2adults?sort=price_a&fs=legdur=-{flight_hours * 60}{stops_param};virtualinterline=-virtualinterline;airportchange=-airportchange"
        if country in ['USA', 'Canada'] and departure_airport_optional and arrival_airport_optional:
            url = f"{base_url}/{departure_airport}-{arrival_airport}/{date_from_str}/{departure_airport_optional}-{arrival_airport_optional}/{date_to_str}/2adults?sort=price_a&fs=legdur=-{flight_hours * 60}{stops_param};virtualinterline=-virtualinterline;airportchange=-airportchange"

        logger.info(f"[Thread {threading.get_ident()}] Accessing: {url}")
        driver.get(url)
        random_delay(1, 2)
        human_like_interaction(driver)

        blocked, driver = handle_possible_blocking(driver, url)
        if blocked:
            logger.info(f"[Thread {threading.get_ident()}] 🔄 Retrying after block resolution for {date_from_str}...")
            random_delay(18, 22)

        load_timeout = random.randint(45, 48)
        progress_bar_xpath_loaded = "//div[@class='skp2 skp2-hidden skp2-inlined' and @role='progressbar']"

        try:
            WebDriverWait(driver, load_timeout).until(
                EC.presence_of_element_located((By.XPATH, progress_bar_xpath_loaded)))
            logger.info(f"[Thread {threading.get_ident()}] ✅ Progress bar hidden - page loaded.")
            time.sleep(1)

            try:
                no_results = WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='c8MCw-header-text' and contains(text(), 'No matching results found.')]")))
                if no_results:
                    logger.info(f"[Thread {threading.get_ident()}] ⚠️ No flights available for {date_from_str}. Skipping.")
                    no_result_dates.add(start_date)
                    try:
                        driver_queue.put(driver)
                    except queue.Full as e:
                        logger.warning(f"⚠️ Driver queue full: {str(e).splitlines()[0]}")
                        driver.quit()
                    except Exception as e:
                        logger.warning(f"⚠️ Failed to return driver to queue: {str(e).splitlines()[0]}")
                        driver.quit()
                    return
            except TimeoutException:
                pass
        except TimeoutException:
            logger.info(f"[Thread {threading.get_ident()}] ⚠️ Timeout waiting for progress bar to be hidden.")
            driver.quit()
            return

        for _ in range(2):
            scroll_amount = random.randint(200, 500)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            random_delay(1, 2)

        flights = driver.find_elements(By.XPATH, "//div[contains(@class, 'nrc6')]")
        if not flights:
            logger.info(f"[Thread {threading.get_ident()}] 🤷🏻‍♂️ No flights found on page for {date_from_str}")
            no_result_dates.add(start_date)
            try:
                driver_queue.put(driver)
            except queue.Full as e:
                logger.warning(f"⚠️ Driver queue full: {str(e).splitlines()[0]}")
                driver.quit()
            except Exception as e:
                logger.warning(f"⚠️ Failed to return driver to queue: {str(e).splitlines()[0]}")
                driver.quit()
            return

        flight = flights[0]
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth'});", flight)
        random_delay(1, 2)

        def safe_extract(xpath, default="-"):
            try:
                text = flight.find_element(By.XPATH, xpath).text.strip()
                return text.replace('$', '').replace(',', '').strip()
            except NoSuchElementException:
                return default

        duration_elements = flight.find_elements(By.XPATH, ".//div[contains(@class, 'xdW8')]/div[contains(@class, 'vmXl')]")
        duration1 = duration_elements[0].text.strip() if len(duration_elements) > 0 else "-"
        duration2 = duration_elements[1].text.strip() if len(duration_elements) > 1 else "-"

        month_name = start_date.strftime('%B')
        formatted_month = month_name[:3]
        formatted_date = f"{start_date.day:02d}-{formatted_month}-{str(start_date.year)[-2:]}"

        price_xpath = ".//div[contains(@class, 'e2GB-price-text-container')]/div[contains(@class, 'e2GB-price-text')]"
        airline_xpath = ".//div[contains(@class, 'J0g6-operator-text')]"
        if country == 'USA':
            airline_xpath = ".//div[contains(@class, 'c_cgF c_cgF-mod-variant-default')]"

        price_text = safe_extract(price_xpath)
        if country == 'Canada':
            price_text = price_text.replace('C ', '')

        if not price_text or not price_text.replace('.', '').isdigit():
            raise ValueError(f"Invalid or missing price: '{price_text}'")

        airline = safe_extract(airline_xpath)

        excel_arrival_airport = arrival_airport
        if departure_airport_optional and arrival_airport_optional:
            excel_arrival_airport = f"{arrival_airport} x {departure_airport_optional}"

        flight_data = {
            'Date': formatted_date,
            'Departure Airport': departure_airport,
            'Arrival Airport': excel_arrival_airport,
            'Nights': nights,
            'Airline': airline,
            'Price': float(price_text),
            'Departure Time': duration1,
            'Arrival Time': duration2
        }

        logger.info(f"[Thread {threading.get_ident()}] 🔍 Found flight: {flight_data['Airline']} for {flight_data['Price']} on {formatted_date}")
        results_queue.put(flight_data)
        driver_queue.put(driver)

    except Exception as e:
        logger.warning(f"[Thread {threading.get_ident()}] ⚠️ Error scraping interval starting {start_date}: {str(e).splitlines()[0]}")
        if driver:
            driver.quit()


@app.route('/', methods=['GET', 'POST'])
def index():
    usa_airports = ["JFK", "EWR", "BOS", "MIA", "MCO", "ORD", "IAH", "IAD", "DEN", "DTW", "PHL", "LAS", "LAX", "SFO", "ATL", "DFW", "SWF"]
    canada_airports = ["YYZ", "YVR", "YOW", "YUL", "YHZ", "YEG", "YYC"]
    selected_country = request.form.get('country')

    if request.method == 'POST':
        departure_airport = request.form['departure_airport']
        arrival_airport = request.form['arrival_airport']
        date_from_str = request.form['date_from']
        date_to_str = request.form['date_to']
        nights = int(request.form['nights'])
        stops = request.form.getlist('stops')
        flight_hours = int(request.form['flight_hours'])
        country = request.form.get('country')
        departure_airport_optional = request.form.get('departure_airport_optional')
        arrival_airport_optional = request.form.get('arrival_airport_optional')

        search_params = {
            'departure_airport': departure_airport,
            'arrival_airport': arrival_airport,
            'date_from': date_from_str,
            'date_to': date_to_str,
            'nights': nights,
            'stops': stops,
            'flight_hours': flight_hours,
            'country': country,
            'departure_airport_optional': departure_airport_optional,
            'arrival_airport_optional': arrival_airport_optional
        }

        logger.info(f"Form Data: {search_params}")

        all_flights = []
        start_date = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        end_date_user = datetime.strptime(date_to_str, '%Y-%m-%d').date()

        interval_starts = []
        current_date = start_date
        while current_date <= end_date_user:
            interval_starts.append(current_date)
            current_date += timedelta(days=1)

        num_threads = min(int(request.form.get('num_tabs', 5)), len(interval_starts))
        driver_queue = Queue(maxsize=num_threads)
        for _ in range(num_threads):
            driver_queue.put(setup_driver())
        results_queue = Queue()
        threads = []

        for start_interval in interval_starts:
            thread = threading.Thread(target=scrape_flight_data_interval, args=(driver_queue, results_queue, search_params, start_interval))
            threads.append(thread)
            threads[-1].start()
            while len(threading.enumerate()) - threading.active_count() > num_threads + 1:
                time.sleep(0.1)

        for thread in threads:
            thread.join()

        while not results_queue.empty():
            flight_data = results_queue.get()
            if flight_data:
                all_flights.append(flight_data)

        while not driver_queue.empty():
            try:
                driver = driver_queue.get()
                # logger.info(f"[Thread {threading.get_ident()}] 🔁 Re-scraping for {start_date}")
                driver.quit()
            except Exception as e:
                logger.warning(f"⚠️ Error quitting driver: {str(e).splitlines()[0]}")

        logger.info(f"✈️ Total number of flights found across all intervals: {len(all_flights)}")

        if all_flights:
            df = pd.DataFrame(all_flights)
            df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%y')
            output_file = f"{departure_airport.upper()} - {arrival_airport.upper()}{f' x {departure_airport_optional.upper()}' if 'departure_airport_optional' in locals() and departure_airport_optional and 'arrival_airport_optional' in locals() and arrival_airport_optional else ''}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            wb = Workbook()
            ws = wb.active
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
                ws.append(row)
                for c_idx, cell in enumerate(ws[r_idx + 1]):
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    if r_idx == 0:
                        cell.font = Font(bold=True)

            date_column = ws['A']
            for cell in date_column:
                cell.number_format = 'DD-MMM-YY'

            wb.save(output_file)
            logger.info("⏳ Waiting briefly to ensure all threads completed writing...")
            time.sleep(2)

            scraped_df = pd.read_excel(output_file, engine='openpyxl')
            logger.info("📄 Data currently in Excel:")
            logger.info(scraped_df[['Date', 'Airline', 'Price']].head(5))

            scraped_dates = set(scraped_df['Date'].dt.date)
            expected_dates = set(interval_starts)
            logger.info(f"📅 Expected dates: {expected_dates}")
            logger.info(f"📅 Scraped dates: {scraped_dates}")

            valid_expected_dates = expected_dates - no_result_dates
            missing_dates = sorted(list(valid_expected_dates - scraped_dates))
            logger.info(f"❗ Final missing: {missing_dates}")

            if missing_dates:
                logger.info(f"🔁 Missing dates detected: {len(missing_dates)}. Re-scraping...")
                logger.info(f"🔁 Missing dates detected: {missing_dates}")
                logger.info("🔄 Starting retry scraping threads...")

                driver_queue = Queue(maxsize=min(num_threads, len(missing_dates)))
                driver_count = min(num_threads, len(missing_dates))

                for i in range(driver_count):
                    try:
                        driver = setup_driver()
                        driver_queue.put(driver)
                        time.sleep(2)
                    except Exception as e:
                        logger.warning(f"❌ Retry driver {i + 1} failed to start: {str(e).splitlines()[0]}")

                results_queue = Queue()
                threads = []

                for retry_date in missing_dates:
                    logger.info(f"[Thread {threading.get_ident()}] 🔁 Re-scraping for {start_date}")
                    thread = threading.Thread(target=scrape_flight_data_interval, args=(driver_queue, results_queue, search_params, retry_date))
                    threads.append(thread)
                    thread.start()

                for thread in threads:
                    thread.join(timeout=90)
                    logger.info("✅ All retry threads completed.")
                retry_flights = []
                while not results_queue.empty():
                    retry_flights.append(results_queue.get())

                if retry_flights:
                    retry_df = pd.DataFrame(retry_flights)
                    retry_df['Date'] = pd.to_datetime(retry_df['Date'], format='%d-%b-%y')
                    df = pd.concat([df, retry_df], ignore_index=True).drop_duplicates(subset=['Date'])

                    wb = Workbook()
                    ws = wb.active
                    for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
                        ws.append(row)
                        for c_idx, cell in enumerate(ws[r_idx + 1]):
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                            if r_idx == 0:
                                cell.font = Font(bold=True)
                    for cell in ws['A']:
                        cell.number_format = 'DD-MMM-YY'
                    wb.save(output_file)
                    logger.info(f"📦 File saved after re-scraping all dates: {output_file}")
                    while not driver_queue.empty():
                        try:
                            driver = driver_queue.get()
                            driver.quit()
                        except Exception as e:
                            logger.warning(f"⚠️ Error quitting retry driver: {str(e).splitlines()[0]}")
            logger.info(f"💾 Final results Saved to {output_file}")
            logger.info("➡️ Rendering results page with final output.")
            if no_result_dates:
                logger.info("📭 Dates with no matching results:")
                for d in sorted(no_result_dates):
                    logger.info(f"❌ {d.strftime('%d-%b-%Y')}")
            return render_template('results.html', output_file=output_file)
        else:
            return render_template('results.html')

    return render_template('index.html', usa_airports=usa_airports, canada_airports=canada_airports, selected_country=selected_country)


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join(str(app.config['UPLOAD_FOLDER']), str(filename))
    try:
        return send_file(filepath, as_attachment=True, download_name=filename)
    except FileNotFoundError:
        return "Error: File not found.", 404


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
