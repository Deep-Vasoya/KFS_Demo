import sys
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime, timedelta
import random
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (NoSuchElementException,
                                        TimeoutException,
                                        WebDriverException)
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side


def random_delay(min_sec=1, max_sec=2):
    """Random delay between actions"""
    time.sleep(random.uniform(min_sec, max_sec))


def human_like_interaction(driver):
    """Simulate human mouse movements and pauses"""
    try:
        action = ActionChains(driver)
        for _ in range(random.randint(1, 5)):
            x = random.randint(-50, 50)
            y = random.randint(-50, 50)
            action.move_by_offset(x, y).perform()
            time.sleep(random.uniform(0.1, 0.3))

        if random.random() > 0.3:
            scroll_amount = random.randint(200, 600)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            time.sleep(random.uniform(0.5, 1.5))
    except Exception as e:
        print(f"Interaction simulation failed: {str(e)}")


def setup_driver():
    """Configure Chrome with maximum stealth settings"""
    options = uc.ChromeOptions()
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument("--disable-notifications")

    user_agents = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36"
    ]
    options.add_argument(f"user-agent={random.choice(user_agents)}")
    options.add_argument(f"--window-size={random.randint(1000, 1400)},{random.randint(800, 1200)}")

    try:
        driver = uc.Chrome(
            options=options,
            headless=False,
            use_subprocess=True
        )

        # Remove webdriver flag and add fake plugins
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                window.navigator.chrome = {
                    runtime: {},
                };
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3]
                });
            """
        })
        return driver
    except Exception as e:
        print(f"❌ Failed to initialize WebDriver: {str(e)}")
        sys.exit(1)


def handle_possible_blocking(driver, current_url):
    """Enhanced blocking handler with cookie clearing and retry"""
    blocking_indicators = [
        "//div[contains(text(), 'Access Denied')]",
        "//div[contains(text(), 'Checking your browser')]",
        "//div[contains(text(), 'Please verify you are a human')]",
        "//iframe[contains(@title, 'recaptcha')]",
        "//div[contains(@class, 'cf-challenge')]"
    ]

    for indicator in blocking_indicators:
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, indicator)))
            print("⚠ Human verification detected. Clearing cookies and retrying after 1 minute...")

            # Clear cookies and storage
            driver.delete_all_cookies()
            driver.execute_script("window.localStorage.clear();")
            driver.execute_script("window.sessionStorage.clear();")

            # Wait for 1 minute
            time.sleep(60)

            # Recreate a driver with a fresh session
            driver.quit()
            driver = setup_driver()

            # Retry the request
            driver.get(current_url)
            human_like_interaction(driver)
            random_delay(1, 2)

            return True, driver
        except:
            continue

    return False, driver


def scrape_flight_data(driver, date_tuple):
    """Scrape flight data with blocking recovery"""
    departure_airport = "JFK"
    arrival_airport = "OSL"
    departure_date_str, return_date_str = date_tuple
    departure_date = datetime.strptime(departure_date_str, '%Y-%m-%d').date()
    return_date = datetime.strptime(return_date_str, '%Y-%m-%d').date()

    nights = (return_date - departure_date).days

    url = f"https://www.kayak.com/flights/{departure_airport}-{arrival_airport}/{departure_date_str}/{return_date_str}/2adults?ucs=3crokd&sort=price_a&fs=legdur=-1200;stops=-2;virtualinterline=-virtualinterline;airportchange=-airportchange"
    print(f"🌐 Accessing: {url}")

    try:
        driver.get(url)
        random_delay(1, 2)
        human_like_interaction(driver)

        # Check for blocking
        blocked, driver = handle_possible_blocking(driver, url)
        if blocked:
            print("🔄 Retrying after block resolution...")
            random_delay(5, 10)

        # Wait for results
        try:
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'nrc6')]")))
        except TimeoutException:
            try:
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//div[contains(text(), 'prices')]")))
            except:
                print("❌ Timed out waiting for flight results")
                return None, driver

        # Human-like scrolling
        for _ in range(3):
            scroll_amount = random.randint(300, 700)
            driver.execute_script(f"window.scrollBy(0, {scroll_amount})")
            random_delay(1, 2)

        flights = driver.find_elements(By.XPATH, "//div[contains(@class, 'nrc6')]")
        if not flights:
            print("❌ No flights found on page")
            return None, driver

        flight = flights[0]
        driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth'});", flight)
        random_delay(1, 2)

        def safe_extract(xpath, default="Unknown"):
            """Try extracting text from an element, return default if missing"""
            try:
                text = flight.find_element(By.XPATH, xpath).text.strip()
                return text.replace('$', '').strip()
            except:
                return default

        duration_elements = flight.find_elements(By.XPATH,
                                                 ".//div[contains(@class, 'xdW8')]/div[contains(@class, 'vmXl')]")
        duration1 = duration_elements[0].text.strip() if len(duration_elements) > 0 else "Unknown"
        duration2 = duration_elements[1].text.strip() if len(duration_elements) > 1 else "Unknown"

        month_name = departure_date.strftime('%B')  # MARK: Get the full month name
        formatted_month = month_name[:3]  # MARK: Take first 3 letters of month
        formatted_date = f"{departure_date.day:02d}-{formatted_month}-{str(departure_date.year)[-2:]}"

        flight_data = {
            'Date': formatted_date,
            'Departure Airport': departure_airport,
            'Arrival Airport': arrival_airport,
            'Nights': nights,
            'Airline': safe_extract(".//div[contains(@class, 'c_cgF')]"),
            'Price': float(safe_extract(
                ".//div[contains(@class, 'e2GB-price-text-container')]/div[contains(@class, 'e2GB-price-text')]")),
            'Departure Time': duration1,
            'Arrival Time': duration2
        }

        print(f"✅ Found flight: {flight_data['Airline']} for {flight_data['Price']}")
        return flight_data, driver

    except Exception as e:
        print(f"❌ Error scraping {departure_date_str}: {str(e)}")
        return None, driver


def main():
    driver = setup_driver()
    all_flights = []

    # Set a search date range
    start_date = datetime(2025, 5, 1)
    end_date = datetime(2025, 7, 31)

    try:
        current_date = start_date
        while current_date <= end_date:
            delay = random.uniform(1.5, 4.5)
            print(f"⏳ Waiting {delay:.1f} seconds before next search...")
            time.sleep(delay)

            # ✅ Change duration here to affect the number of nights
            return_date = current_date + timedelta(days=7)
            date_tuple = (current_date.strftime('%Y-%m-%d'), return_date.strftime('%Y-%m-%d'))

            flight_data, driver = scrape_flight_data(driver, date_tuple)
            if flight_data:
                all_flights.append(flight_data)

            current_date += timedelta(days=1)

            if random.random() > 0.7:
                human_like_interaction(driver)

    except KeyboardInterrupt:
        print("⏹ Script interrupted by user")
    except Exception as e:
        print(f"❌ Fatal error: {str(e)}")
    finally:
        if all_flights:
            # ✅ Save to Excel using openpyxl for formatting
            df = pd.DataFrame(all_flights)
            df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%y')
            output_file = "kayak_flights.xlsx"

            wb = Workbook()
            ws = wb.active

            # Define border style
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for r_idx, row in enumerate(dataframe_to_rows(df, header=True, index=False)):
                ws.append(row)
                for c_idx, cell in enumerate(ws[r_idx + 1]):
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                    if r_idx == 0:
                        cell.font = Font(bold=True)

            # Format the Date column to display as DD-MMM-YY
            date_column = ws['A']
            for cell in date_column:
                cell.number_format = 'DD-MMM-YY'

            wb.save(output_file)
            print(f"💾 Saved results to {output_file}")

        try:
            driver.quit()
        except:
            pass
        sys.exit(0)


if __name__ == "__main__":
    main()
